from openpyxl import load_workbook

load_wb = load_workbook("C:\\Users\\twins\\OneDrive\\문서\\카카오톡 받은 파일\\밀화부리\\스캔본.xlsx", data_only=True)
load_ws = load_wb['Sheet1']

scanlist = []
for i in range(2, 42):
    scanlist.append(load_ws.cell(i, 4).value)
print(scanlist)

realList = {}
noList = {}

for i in range(2, 2179):
    realList[load_ws.cell(i, 6).value] = []
    noList[load_ws.cell(i, 6).value] = []

# print(realList)

for i in range(2, load_ws.max_row + 1):
    realList[load_ws.cell(i, 2).value].append(load_ws.cell(i, 1).value)

for _id in realList.keys():
    for value in scanlist:

        if value not in realList[_id]:
            noList[_id].append(value)
    print(noList[_id])

write_ws = load_wb.create_sheet('생성시트')
i = 1
for _id in noList.keys():

    for value in noList[_id]:
        write_ws.cell(i, 1).value = _id
        write_ws.cell(i, 2).value = value
        i += 1
load_wb.save("C:\\Users\\twins\\OneDrive\\문서\\카카오톡 받은 파일\\밀화부리\\스캔본1.xlsx")
