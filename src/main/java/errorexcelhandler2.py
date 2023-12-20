from openpyxl import load_workbook

load_wb = load_workbook("C:\\Users\\twins\\OneDrive\\문서\\카카오톡 받은 파일\\찌르레기\\TOTAL_ERROR_NO_SCAN.xlsx",
                        data_only=True)

scanList = {
    "c_doc1y": "1세 증상",
    "c_doc2y": "2세 증상", "c_doc3y": "3세 증상", "c_doc4y": "4세 증상", "c_doc5y": "5세 증상", "c_doc6y": "6세 증상",
    "c_doc7y": "7세 증상", "c_doc8y": "8세 증상", "c_doc9y": "9세 증상", "c_doc6m": "6개월 증상",
    "c_en1y": "1세 환경", "c_en2y": "2세 환경", "c_en3y": "3세 환경", "c_en4y": "4세 환경", "c_en5y": "5세 환경",
    "c_en6y": "6세 환경", "c_en7y": "7세 환경", "c_en8y": "8세 환경", "c_en9y": "9세 환경", "m_en1y": "출산전 환경",
    "c_en6m": "6개월 환경",
    "c_fu1y": "1세 추적", "c_fu2y": "2세 추적", "c_fu3y": "3세 추적", "c_fu4y": "4세 추적", "c_fu5y": "5세 추적",
    "c_fu6y": "6세 추적", "c_fu7y": "7세 추적", "c_fu8y": "8세 추적", "c_fu9y": "9세 추적",
    "c_fu6m": "6개월 추적",
}

total_idList = {}
total_load_ws = load_wb['total no scan']
for i in range(2, total_load_ws.max_row + 1):
    total_idList[total_load_ws.cell(i, 1).value] = set([])

for i in range(2, total_load_ws.max_row + 1):
    for key in scanList.keys():
        if key in total_load_ws.cell(i, 2).value:
            total_idList[total_load_ws.cell(i, 1).value].add(key)
            break

load_ws = load_wb['error']
for i in range(2, load_ws.max_row + 1):

    now_id = load_ws.cell(i, 1).value
    now_val = load_ws.cell(i, 2).value

    if now_id not in total_idList.keys():
        load_ws.cell(i, 5).value = 0

        continue

    now_survey_list = total_idList[now_id]

    if len(now_survey_list) == 0:
        load_ws.cell(i, 5).value = 1
        continue

    for now_survey in now_survey_list:

        if now_survey in now_val:
            load_ws.cell(i, 5).value = 0

            break
        if now_survey not in now_val:

            load_ws.cell(i, 5).value = 1

            continue
        else:

            load_ws.cell(i, 5).value = 2



load_wb.save("C:\\Users\\twins\\OneDrive\\문서\\카카오톡 받은 파일\\찌르레기\\TOTAL_ERROR_NO_SCAN.xlsx")
