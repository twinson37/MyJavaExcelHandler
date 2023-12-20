from openpyxl import load_workbook

load_wb = load_workbook("스캔 파일 없는 결측 값 수정로그.xlsx", data_only=True)
load_sheet = load_wb["스캔본 목록"]

for i in range(2, load_sheet.max_row+1):
    file_name = load_sheet.cell(i, 2).value
    survey_name = load_sheet.cell(i, 4).value
    if survey_name in file_name:
        continue
    elif survey_name not in file_name:
        load_sheet.cell(i, 5).value = "no"
    else:
        load_sheet.cell(i, 5).value = "check"

load_wb.save("스캔 파일 없는 결측 값 수정로그(수정1).xlsx")
