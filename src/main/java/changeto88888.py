from openpyxl import load_workbook

data_wb = load_workbook(
    'C:\\Users\\twins\\OneDrive\\문서\\카카오톡 받은 파일\\찌르레기\\final_acy_통합_23.12.18.xlsx',
    data_only=True)

data_ws = data_wb['Sheet1']
no_scan_ws = data_wb['Sheet2']

idList = {}

for i in range(2, data_ws.max_row + 1):
    _id = data_ws.cell(i, 1).value
    idList[_id] = i

varList = {}

for i in range(1, data_ws.max_column + 1):
    var_id = data_ws.cell(1, i).value
    varList[str(var_id).lower()] = i

print(no_scan_ws.max_row+1)
for i in range(2, no_scan_ws.max_row+1):
    now_id = no_scan_ws.cell(i, 1).value
    now_var = str(no_scan_ws.cell(i, 2).value).lower()
    # print(i,now_var)

    now_id_index = idList[now_id]
    now_var_index = varList[now_var]

    data_ws.cell(now_id_index, now_var_index).value = 88888
    no_scan_ws.cell(i, 6).value = 88888

data_wb.save("C:\\Users\\twins\\OneDrive\\문서\\카카오톡 받은 파일\\찌르레기\\final_acy_통합_23.12.18(88888 수정).xlsx")

