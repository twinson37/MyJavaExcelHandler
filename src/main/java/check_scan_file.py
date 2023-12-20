from openpyxl import load_workbook

load_wb = load_workbook("스캔 파일 없는 결측 값 수정로그(수정1).xlsx", data_only=True)
load_sheet = load_wb["스캔본 목록"]
write_sheet = load_wb["샘플ID_스캔본 존재여부"]

var_index = {}

for i in range(2, write_sheet.max_column + 1):
    var_name = write_sheet.cell(1, i).value
    var_index[var_name] = i

id_index = {}

for i in range(2, write_sheet.max_row + 1):
    now_id = write_sheet.cell(i, 1).value
    id_index[now_id] = i

for i in range(2, load_sheet.max_row + 1):
    now_id = load_sheet.cell(i, 1).value
    now_var = load_sheet.cell(i, 4).value
    if now_var not in var_index.keys():
        continue
    now_var_index = var_index[now_var]
    if now_id not in id_index.keys():
        continue
    now_id_index = id_index[now_id]

    write_sheet.cell(now_id_index, now_var_index).value = 1

for i in range(2, write_sheet.max_row + 1):
    for j in range(2, write_sheet.max_column + 1):
        if write_sheet.cell(i, j).value != 1:
            write_sheet.cell(i, j).value = 0

load_wb.save("스캔 파일 없는 결측 값 수정로그(final).xlsx")