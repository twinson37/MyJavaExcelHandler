from openpyxl import load_workbook

codebook_wb = load_workbook("C:\\Users\\twins\\OneDrive\\문서\\카카오톡 받은 파일\\standard_codebook_total_23.12.04.xlsx",
                            data_only=True)
_7y_data_wb = load_workbook(
    'C:\\Users\\twins\\OneDrive\\문서\\카카오톡 받은 파일\\밀화부리\\final_acy9y_permittion_23.12.05-all.xlsx',
    data_only=True)

_7y_data_ws = _7y_data_wb['Sheet1']
codebook_ws = codebook_wb['Sheet1']
number_scan = codebook_wb['Sheet2']
no_scan_ws = codebook_wb['Sheet3']

table_to_number = {}
number_to_table = {}

for i in range(2, number_scan.max_row + 1):
    table = number_scan.cell(i, 1).value
    number = str(number_scan.cell(i, 2).value)

    table_to_number[table] = number
    number_to_table[number] = table
# print(number_scan_dict)

no_scan_list = {}
for i in range(1, no_scan_ws.max_row + 1):
    _id = no_scan_ws.cell(i, 1).value
    number = str(no_scan_ws.cell(i, 2).value)
    # print(number)
    if number not in no_scan_list.keys():
        no_scan_list[number] = []
    no_scan_list[number].append(_id)

# print(no_scan_list)

var_table_relation = {}
var_catogory_relation = {}
for i in range(2, codebook_ws.max_row + 1):
    var_name = codebook_ws.cell(i, 8).value
    table_name = codebook_ws.cell(i, 2).value
    var_category = codebook_ws.cell(i, 12).value
    var_table_relation[var_name] = table_name
    if var_category is None:
        var_category = ""
    var_catogory_relation[var_name] = var_category[:1]
    # if var_category is None:
    #     var_catogory_relation[var_name] = ""
    # else:
    #     var_catogory_relation[var_name] = var_category[:1]
# print(var_table_relation)
# id와 행번호
id_row_number = {}

for i in range(2, _7y_data_ws.max_row + 1):
    _id = _7y_data_ws.cell(i, 1).value
    id_row_number[_id] = i
print(id_row_number)

errorList = []
for i in range(1, _7y_data_ws.max_column + 1):
    now_var_name = _7y_data_ws.cell(1, i).value
    print(now_var_name)
    if now_var_name not in var_table_relation.keys():
        print(now_var_name, " not in var_table_relation")
        continue
    now_table = var_table_relation[now_var_name]
    if now_table not in table_to_number.keys():
        print(now_table, "not in table_to_number")
        continue
    nowNumber = table_to_number[now_table]

    if nowNumber not in no_scan_list.keys():
        print(nowNumber, " not in no_scan_list")
        continue
    now_no_scan__list = no_scan_list[nowNumber]
    for _id in no_scan_list[nowNumber]:
        if _id not in id_row_number:
            print(_id, " not in id_row_number")

            continue
        now_row_num = id_row_number[_id]
        print(_id, " ", now_var_name, " ", _7y_data_ws.cell(now_row_num, i).value)
        if (_7y_data_ws.cell(now_row_num, i).value == "66666"
                or _7y_data_ws.cell(now_row_num, i).value == "77777"
                or _7y_data_ws.cell(now_row_num, i).value == "99999"
                or _7y_data_ws.cell(now_row_num, i).value == 66666
                or _7y_data_ws.cell(now_row_num, i).value == 99999
                or _7y_data_ws.cell(now_row_num, i).value == 777777):

            if var_catogory_relation[now_var_name] != "":
                errorList.append(
                    [_id, now_var_name, _7y_data_ws.cell(now_row_num, i).value, var_catogory_relation[now_var_name]]
                )
                _7y_data_ws.cell(now_row_num, i).value = var_catogory_relation[now_var_name]
            else:
                errorList.append(
                    [_id, now_var_name, _7y_data_ws.cell(now_row_num, i).value, "미수정"]
                )

print(errorList)
codebook_write_ws = codebook_wb.create_sheet("correct_log")

for error in errorList:
    print(error)
    codebook_write_ws.append(error)

codebook_wb.save("C:\\Users\\twins\\OneDrive\\문서\\카카오톡 받은 파일\\밀화부리\\new_코드북_9Y.xlsx")
_7y_data_wb.save("final_acy9y_permittion_23.12.05-all).xlsx")