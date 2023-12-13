from queue import Queue
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
import time

global relation
global data_sheet
global data_var_list
global write_ws
global data_work_book
global data_sheet2
global data_work_book2
global excepted
global data_filename
global codebook_filename


class VarNode:

    def __init__(self, var_name, var_prior, var_prior_add_val, column_number):
        self.var_name = var_name
        self.var_prior = var_prior
        self.var_prior_add_val = var_prior_add_val
        self.column_number = column_number


def load_data():
    start = time.time()
    global data_sheet
    global data_work_book
    global data_sheet2
    global data_work_book2

    print("데이터 로드 중..")
    data_work_book = load_workbook(data_filename,
                                   read_only=False,
                                   data_only=False)
    data_sheet = data_work_book["Sheet1"]
    data_work_book2 = load_workbook(data_filename,
                                    read_only=False,
                                    data_only=False)
    data_sheet2 = data_work_book2["Sheet1"]
    print("time :", time.time() - start, "sec..done")


def setRelation():
    start = time.time()
    global relation
    global data_var_list
    global excepted

    purpleFill = PatternFill(start_color='DFCDFF',
                             end_color='DFCDFF',
                             fill_type='solid')
    excepted = []
    print("relation 설정 중..")
    relation = {}
    load_work_book = load_workbook(codebook_filename,
                                   read_only=False,
                                   data_only=False)

    load_sheet = load_work_book["Sheet1"]
    data_var_list = [""]

    for i in range(1, data_sheet.max_column + 1):
        data_var_list.append(data_sheet.cell(1, i).value)

    for i in range(1, load_sheet.max_row + 1):

        if load_sheet.cell(i, 16).value == 1:

            name = load_sheet.cell(i, 8).value

            if name not in data_var_list:
                excepted.append(name)
                continue

            var_prior = load_sheet.cell(i, 17).value
            var_prior_val = load_sheet.cell(i, 18).value
            column_index = data_var_list.index(name)

            if var_prior not in relation:
                if var_prior not in data_var_list:
                    excepted.append(var_prior)
                    continue

                index = data_var_list.index(var_prior)
                relation[var_prior] = [index, []]

            now_var = VarNode(name, var_prior, var_prior_val, column_index)
            relation.get(var_prior)[1].append(now_var)
            data_sheet.cell(1, column_index).fill = purpleFill
            data_sheet.cell(1, data_var_list.index(var_prior)).fill = purpleFill
            data_sheet2.cell(1, column_index).fill = purpleFill
            data_sheet2.cell(1, data_var_list.index(var_prior)).fill = purpleFill

    print("data_var_list : ", data_var_list)
    print("relation : ", relation)
    print("time :", time.time() - start, "sec..done")


def bfs():
    global relation
    start = time.time()
    print("탐색 중..")
    visited = [0] * len(data_var_list)

    for key in relation.keys():

        key_index = relation[key][0]

        if visited[key_index] == 1:
            continue

        visited[key_index] = 1
        queue = Queue()
        queue.put(key)

        while queue.qsize() != 0:

            now_name = queue.get()

            if now_name in relation:

                for node in relation.get(now_name)[1]:
                    next_index = node.column_number

                    if visited[next_index] == 0:
                        visited[next_index] = 1
                        queue.put(node.var_name)
                        print(now_name, ": ", node.var_name)
                        find_Error(now_name, node)

    print("time :", time.time() - start, "sec..done")


def find_Error(parent_name, child_node):
    global relation
    global data_sheet
    global data_var_list
    global write_ws

    parent_index = data_var_list.index(parent_name)
    child_index = data_var_list.index(child_node.var_name)
    var_name = child_node.var_name
    var_prior = parent_name
    var_prior_add_val = str(child_node.var_prior_add_val)

    yellowFill = PatternFill(start_color='FFFFFF00',
                             end_color='FFFFFF00',
                             fill_type='solid')

    pinkFill = PatternFill(start_color='F3CDFF',
                           end_color='F3CDFF',
                           fill_type='solid')

    isError = 0
    for i in range(2, data_sheet.max_row + 1):

        sample_id = data_sheet.cell(i, 1).value
        value = str(data_sheet.cell(i, child_index).value)
        var_prior_val = str(data_sheet.cell(i, parent_index).value)

        if ((var_prior_val is None or var_prior_val == "" or var_prior_val == "None")
                and (value == 77777 or value == "77777"
                     or value == "66666" or value == "99999"
                     or value == 66666 or value == 99999)):
            write_ws.append([sample_id, var_name, value, var_prior, var_prior_val, var_prior_add_val, 1])
            data_sheet2.cell(i, child_index).value = "BLANK!"
            data_sheet.cell(1, child_index).fill = pinkFill
            data_sheet2.cell(1, child_index).fill = pinkFill

            isError = 1

        if (var_prior_val == var_prior_add_val
                and (value == 77777 or value == "77777")):
            write_ws.append([sample_id, var_name, value, var_prior, var_prior_val, var_prior_add_val, 2])
            data_sheet2.cell(i, child_index).value = "0"
            data_sheet2.cell(1, child_index).fill = pinkFill

            isError = 1

    if isError == 1:
        data_sheet.cell(1, parent_index).fill = yellowFill
        data_sheet2.cell(1, parent_index).fill = yellowFill


def init():
    load_data()
    setRelation()


def make_File():
    global data_filename, codebook_filename, write_ws

    init()
    write_wb = Workbook()
    write_ws = write_wb.active
    write_ws.append(["ID", "var_name", " value", "var_prior", "var_prior_val", "var_prior_add_val", "type"])
    bfs()
    error_sheet = write_wb.create_sheet("제외 또는 직접 조사")
    for excepted_name in excepted:
        error_sheet.append([excepted_name])
    write_wb.save(f"{error_log_filename}{time.time()}.xlsx")


if __name__ == "__main__":

    data_filename = "/Users/kimjungi/Desktop/rex/final/final_acy8y_cocoa_error5_modified_23.12.05.xlsx"
    codebook_filename = "/Users/kimjungi/Desktop/rex/final/8세통합코딩북_23.12.05.xlsx"
    error_log_filename = "/Users/kimjungi/Desktop/rex/final/error_log_8y"
    make_File()

    data_filename = "/Users/kimjungi/Desktop/rex/final/final_acy9y_permittion_23.12.05-all.xlsx"
    codebook_filename = "/Users/kimjungi/Desktop/rex/final/9세 코드북_23.12.05_real.xlsx"
    error_log_filename = "/Users/kimjungi/Desktop/rex/final/error_log_9y"
    make_File()

    data_filename = "/Users/kimjungi/Desktop/rex/final/final_acy07y_23.12.05_spt수정.xlsx"
    codebook_filename = "/Users/kimjungi/Desktop/rex/final/0-7 codebook.xlsx"
    error_log_filename = "/Users/kimjungi/Desktop/rex/final/error_log_7y"
    make_File()

    # data_work_book.save(f"{data_filename}(자동수정)_{time.time()}.xlsx")
    # data_work_book2.save(f"{data_filename}(자동수정)2_{time.time()}.xlsx")
