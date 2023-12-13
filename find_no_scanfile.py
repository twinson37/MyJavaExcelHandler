import datetime
import os
import string
from os import path
from os.path import splitext
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
import time


def time_trace(func):
    def wrapper(*args, **kwargs):
        print(func.__name__, '함수 시작')
        st = time.time()
        rt = func(*args, **kwargs)
        print(f'### {func.__name__}({args}) time : {time.time() - st:.3f}ms')
        return rt

    return wrapper


# class HaveScanFileVar:
#
#     def __init__(self, var_name, var_prior, var_prior_add_val):
#         self.var_name = var_name
#         self.var_prior = var_prior
#         self.var_prior_add_val = var_prior_add_val
#
#
# @time_trace
# def get_var_list():
#     """설문지별 변수 딕셔너리 및 var prior 딕셔너리 생성"""
#     global codebook_sheet, relations, var_prior_relation, survey_name
#     for i in range(1, codebook_sheet.max_row + 1):
#
#         now_survey_name = codebook_sheet.cell(i, 2).value
#         if now_survey_name is not None:
#             now_survey_name = codebook_sheet.cell(i, 2).value.lower()
#         now_var_name = codebook_sheet.cell(i, 8).value
#         if now_var_name is not None:
#             now_var_name = codebook_sheet.cell(i, 8).value.lower()
#         now_var_prior_name = codebook_sheet.cell(i, 17).value
#         if now_var_prior_name is not None:
#             now_var_prior_name = codebook_sheet.cell(i, 17).value.lower()
#
#         if now_survey_name not in survey_name:
#             survey_name[now_survey_name] = []
#         survey_name[now_survey_name].append(now_var_name)
#
#         if now_var_prior_name not in var_prior_relation:
#             var_prior_relation[now_var_prior_name] = []
#         var_prior_relation[now_var_prior_name].append(now_var_name)
#
#     # print(survey_name)
#     # print(var_prior_relation)
#
#     # for relation in relations:
#     #     print(now_var_name[:len(relation)])
#     #     print(relation)
#     #     if now_var_name[:len(relation)] == relation:
#     #         relations[relation].append(now_var_name)
#     #
#     # if now_var_prior_name not in var_prior_relation:
#     #     var_prior_relation[now_var_prior_name] = []
#     # var_prior_relation[now_var_prior_name].append(now_var_name)
#
#     # print(relations)
#
#
# @time_trace
# def make_relation():
#     """설문지 종류 별 리스트 생성"""
#
#     global year_survey_list
#
#     years = {"6개월":"6m", "1세":"1y", "2세":"2y", "3세":"3y", "4세":"4y", "5세":"5y",
#              "6세":"6y", "7세":"7y", "8세":"8y", "9세":"9y"}
#     surveys = {"정신":"ps", "증상":"doc", "추적":"fu", "환경":"en"}
#     year_survey_list["배우자"] = []
#     year_survey_list["출산전_환경"] = []
#
#     for survey in surveys.keys():
#
#         for year in years.keys():
#             year_value = years[year]
#             survey_value = surveys[survey]
#             year_survey_list[year_value + survey_value] = year+survey
#
#     print(year_survey_list)
#
#
# def make_var_prior_relation():
#     pass
#
#
# @time_trace
# def load_data_files():
#     """데이터 파일 불러옴"""
#     now_data_workbooks = []
#     data_paths = ["/Users/kimjungi/Desktop/rex/final/final_acy07y_23.12.05_spt수정.xlsx",
#                   "/Users/kimjungi/Desktop/rex/final/final_acy8y_cocoa_error5_modified_23.12.05.xlsx",
#                   "/Users/kimjungi/Desktop/rex/final/final_acy9y_permittion_23.12.05-all.xlsx"]
#
#     # data_workbooks [[workbook1,workbook1's sheet],...]
#     for data_path in data_paths:
#         now_data_workbook = load_workbook(data_path,
#                                           read_only=False,
#                                           data_only=False)
#         now_data_sheet = now_data_workbook["Sheet1"]
#         now_data_workbooks.append([now_data_workbook, now_data_sheet])
#
#     return now_data_workbooks
#
#
# @time_trace
# def load_codebook():
#     """코드북 불러옴"""
#     # codebook_name, codebook_ext = splitext(codebook_path)
#     codebook_path = "/Users/kimjungi/Desktop/rex/final/standard_codebook_total_23.12.04.xlsx"
#     codebook_workbook = load_workbook(codebook_path,
#                                       read_only=False,
#                                       data_only=False)
#     return codebook_workbook["Sheet1"]
#
#
# def write_workbook():
#     pass
#
#
# def get_data_var_list():
#     global year_survey_list
#     for data_workbook in data_workbooks:
#         for i in range(1, data_workbook.max_column + 1):
#
#             now_col_name = data_workbook.cell(1, i)
#
#             if now_col_name[:1] == 'f':
#                 year_survey_list["배우자"].append(i)
#             if now_col_name[:1] == 'm':
#                 year_survey_list["출산전_환경"].append(i)
#             if now_col_name[:1] == 'c':
#                 pass


if __name__ == "__main__":
    # survey_name = {}
    # relations = {}
    # year_survey_list = {}
    # var_prior_relation = {}
    # data_workbooks = load_data_files()
    # codebook_sheet = load_codebook()
    #
    # # make_relation()
    # # make_var_prior_relation()
    # get_var_list()
    # get_data_var_list()
    # write_workbook()

    scan_files_workbook = load_workbook("/Users/kimjungi/Desktop/rex/바탕화면에 있던 거/스캔본.xlsx")
    scan_files_sheet = scan_files_workbook["Sheet0"]
    # scan_numbers = {"05": "출산전_환경",
    #                 "07": "6개월_증상", "08": "6개월_환경", "09": "6개월_추적",
    #                 "11": "1세_증상", "12": "1세_환경", "13": "1세_추적", "14": "1세_정신",
    #                 "21": "2세_증상", "22": "2세_환경", "23": "2세_추적", "24": "2세_정신",
    #                 "31": "3세_증상", "32": "3세_환경", "33": "3세_추적", "34": "3세_정신",
    #                 "41": "4세_증상", "42": "4세_환경", "43": "4세_추적", "44": "4세_정신",
    #                 "51": "5세_증상", "52": "5세_환경", "53": "5세_추적", "54": "5세_정신",
    #                 "61": "6세_증상", "62": "6세_환경", "63": "6세_추적", "64": "6세_정신",
    #                 "71": "7세_증상", "72": "7세_환경", "73": "7세_추적", "74": "7세_정신",
    #                 "81": "8세_증상", "82": "8세_환경", "83": "8세_추적", "84": "8세_정신",
    #                 "91": "9세_증상", "92": "9세_환경", "93": "9세_추적", "94": "9세_정신",
    #                 }

    scan_names = {"출산전_환경": "05",
                  "6개월_증상": "07", "6개월_환경": "08", "6개월_추적": "09",
                  "1세_증상": "11", "1세_환경": "12", "1세_추적": "13", "1세_정신": "14",
                  "2세_증상": "21", "2세_환경": "22", "2세_추적": "23", "2세_정신": "24",
                  "3세_증상":"31", "3세_환경":"32", "3세_추적":"33", "3세_정신":"34",
                  "41": "4세_증상", "42": "4세_환경", "43": "4세_추적", "44": "4세_정신",
                  "51": "5세_증상", "52": "5세_환경", "53": "5세_추적", "54": "5세_정신",
                  "61": "6세_증상", "62": "6세_환경", "63": "6세_추적", "64": "6세_정신",
                  "71": "7세_증상", "72": "7세_환경", "73": "7세_추적", "74": "7세_정신",
                  "81": "8세_증상", "82": "8세_환경", "83": "8세_추적", "84": "8세_정신",
                  "91": "9세_증상", "92": "9세_환경", "93": "9세_추적", "94": "9세_정신",
                  }
    sample_scan_list = {}
    for i in range(2, scan_files_sheet.max_row + 1):
        sample_id = scan_files_sheet.cell(i, 1).value
        right_under_score_index = scan_files_sheet.cell(i, 2).value.rindex("_")
        left_under_score_index = scan_files_sheet.cell(i, 2).value.index("_")
        scan_file_name = scan_files_sheet.cell(i, 2).value[:right_under_score_index]
        scan_file_number = scan_files_sheet.cell(i, 2).value[:left_under_score_index]
        if scan_file_number in scan_numbers.keys():
            if sample_id not in sample_scan_list:
                sample_scan_list[sample_id] = []
            sample_scan_list[sample_id].append(scan_file_number)

    no_sample_scan_list = {}

    for sample_id in sample_scan_list:
        for scan_number in scan_numbers:
            if scan_number not in sample_scan_list[sample_id]:
                if sample_id not in no_sample_scan_list:
                    no_sample_scan_list[sample_id] = []
                    # print(sample_id, ": ", scan_number)
                no_sample_scan_list[sample_id].append(scan_number)
    print(no_sample_scan_list)

    # if scan_file_number not in scan_numbers.keys() and scan_file_name not in scan_numbers.values():
    #     sample_scan_list[sample_id].append(scan_file_name)

    # for scan_number in scan_numbers.keys():
    #     for sample_scan in sample_scan_list.keys():
    #         file_list = sample_scan_list[sample_scan]
    #
    #         for file in file_list:
    #             if sca

    # print(sample_scan_list)
